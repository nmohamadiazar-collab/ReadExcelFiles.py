import os
from openpyxl import load_workbook
import pandas as pd

# File picker (built-in)
import tkinter as tk
from tkinter import filedialog, messagebox

# ==========================
# SETTINGS (your fixed cells)
# ==========================
START_SHEET_NUMBER = 2   # start from sheet #2
CELL_LABEL = "A21"
CELL_VALUE = "Z46"

def process_one_file(file_path: str) -> pd.DataFrame:
    wb = load_workbook(file_path, data_only=True)
    sheetnames = wb.sheetnames
    total_sheets = len(sheetnames)

    rows = []
    for sheet_number in range(START_SHEET_NUMBER, total_sheets + 1):
        sheet_name = sheetnames[sheet_number - 1]
        ws = wb[sheet_name]

        label = ws[CELL_LABEL].value
        value = ws[CELL_VALUE].value

        rows.append({
            "source_file": os.path.basename(file_path),
            "sheet_number": sheet_number,
            "sheet_name": sheet_name,
            CELL_LABEL: label,
            CELL_VALUE: value
        })

    return pd.DataFrame(rows)

def main():
    # Create OUTPUT folder next to this script
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, "OUTPUT")
    os.makedirs(output_dir, exist_ok=True)

    # Launch file picker
    root = tk.Tk()
    root.withdraw()  # hide the blank Tk window

    file_paths = filedialog.askopenfilenames(
        title="Select Excel file(s) to extract A21 and Z46",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )

    if not file_paths:
        print("No files selected. Exiting.")
        return

    all_results = []

    for file_path in file_paths:
        fname = os.path.basename(file_path)
        print(f"Processing: {fname}")

        try:
            df = process_one_file(file_path)
        except Exception as e:
            print(f"  ERROR reading {fname}: {e}")
            continue

        # Save per-file results
        base_name = os.path.splitext(fname)[0]
        out_xlsx = os.path.join(output_dir, f"{base_name}_A21_Z46.xlsx")
        out_csv = os.path.join(output_dir, f"{base_name}_A21_Z46.csv")

        df.to_excel(out_xlsx, index=False)
        df.to_csv(out_csv, index=False)

        print(f"  Saved: {out_xlsx}")
        all_results.append(df)

    # Save combined results
    if all_results:
        combined = pd.concat(all_results, ignore_index=True)
        combined_xlsx = os.path.join(output_dir, "ALL_SELECTED_FILES_A21_Z46.xlsx")
        combined_csv = os.path.join(output_dir, "ALL_SELECTED_FILES_A21_Z46.csv")
        combined.to_excel(combined_xlsx, index=False)
        combined.to_csv(combined_csv, index=False)

        print("\nDONE.")
        print(f"Combined saved: {combined_xlsx}")
        print(f"Combined saved: {combined_csv}")

        messagebox.showinfo(
            "Done",
            f"Extraction complete!\n\nOutputs saved in:\n{output_dir}"
        )
    else:
        messagebox.showwarning("No output", "No files were processed successfully.")

if __name__ == "__main__":
    main()
