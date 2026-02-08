from openpyxl import load_workbook
import pandas as pd

# =========================================================
# WEEK 1: Read cell Z47 from specific sheet numbers
# and report totals by group
# =========================================================

# -------------------------
# CHANGE THIS ONLY
# -------------------------
FILE_PATH = r"C:\X\Y\Z\X.xlsx"   # <-- put your Week 1 Excel file path here

# -------------------------
# CONFIG (your sheet numbers)
# Sheet numbers are 1-based positions: sheet #1 = first Excel tab
# -------------------------
CELL = "Z47"

GROUPS = {
    "Hiawassee EB": [8, 15, 22, 29, 36, 43, 50, 57, 64],
    "Hiawassee WB": [71, 78, 85, 92, 99, 106, 113, 120, 127],
    "Good Homes WB On Ramp": [134, 141],
    "Good Homes EB Off Ramp": [148, 155],
    "Hiawassee EB On Ramp": [162, 169],
    "Hiawassee WB Off Ramp": [176, 183],
}

def safe_float(x):
    """Convert Excel cell value to float if possible."""
    try:
        if x is None:
            return None
        s = str(x).strip().replace(",", "")
        if s == "":
            return None
        return float(s)
    except Exception:
        return None

def main():
    wb = load_workbook(FILE_PATH, data_only=True)
    sheetnames = wb.sheetnames
    total_sheets = len(sheetnames)

    rows = []

    for group_name, sheet_numbers in GROUPS.items():
        group_sum = 0.0
        group_missing = 0

        for n in sheet_numbers:
            # Validate sheet index
            if n < 1 or n > total_sheets:
                rows.append({
                    "group": group_name,
                    "sheet_number": n,
                    "sheet_name": None,
                    "cell": CELL,
                    "value": None,
                    "note": f"Sheet number out of range (file has {total_sheets} sheets)"
                })
                group_missing += 1
                continue

            sheet_name = sheetnames[n - 1]  # 1-based to 0-based
            ws = wb[sheet_name]

            raw_val = ws[CELL].value
            num_val = safe_float(raw_val)

            if num_val is None:
                group_missing += 1
                note = "Cell empty or not numeric"
            else:
                group_sum += num_val
                note = ""

            rows.append({
                "group": group_name,
                "sheet_number": n,
                "sheet_name": sheet_name,
                "cell": CELL,
                "value": raw_val,
                "note": note
            })

        # Add a summary row for this group
        rows.append({
            "group": group_name,
            "sheet_number": "",
            "sheet_name": "TOTAL (group sum)",
            "cell": CELL,
            "value": group_sum,
            "note": f"missing/invalid sheets: {group_missing}"
        })

    df = pd.DataFrame(rows)

    # Print summary in IDLE shell
    print("\n=== WEEK 1 SUMMARY (cell Z47) ===")
    for group_name in GROUPS.keys():
        total_row = df[(df["group"] == group_name) & (df["sheet_name"] == "TOTAL (group sum)")]
        total_val = total_row["value"].values[0] if not total_row.empty else 0
        print(f"{group_name}: {total_val}")

    # Save detailed logs to files
    df.to_excel("week1_Z47_extraction.xlsx", index=False)
    df.to_csv("week1_Z47_extraction.csv", index=False)

    print("\nSaved: week1_Z47_extraction.xlsx")
    print("Saved: week1_Z47_extraction.csv")

if __name__ == "__main__":
    main()
